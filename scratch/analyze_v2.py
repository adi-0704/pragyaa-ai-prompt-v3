import pandas as pd
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Read using openpyxl to handle merged headers
import openpyxl
wb = openpyxl.load_workbook('AI vs Verifier audits DC.xlsx', data_only=True)

for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"Sheet: {sn} ({ws.max_row} rows x {ws.max_column} cols)")
    
    # Extract all rows as list of dicts using row 1 as headers
    headers = [cell.value for cell in ws[1]]
    print(f"Headers: {headers}")
    print(f"Number of headers: {len(headers)}")
    
    data = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for idx, cell in enumerate(ws[r]):
            if idx < len(headers) and headers[idx]:
                row[headers[idx]] = cell.value
        data.append(row)
    
    df = pd.DataFrame(data)
    print(f"DataFrame columns: {list(df.columns)}")
    print(f"DataFrame shape: {df.shape}")
    
    # Find AI and Verifier columns
    ai_col = None
    ver_col = None
    for c in df.columns:
        cs = str(c).lower()
        if 'call status ai' in cs:
            ai_col = c
        elif 'call status verifier' in cs:
            ver_col = c
    
    print(f"AI col: {ai_col}")
    print(f"Verifier col: {ver_col}")
    
    if ai_col and ver_col:
        df['ai_norm'] = df[ai_col].astype(str).str.strip().str.lower()
        df['ver_norm'] = df[ver_col].astype(str).str.strip().str.lower()
        
        print("\n=== Cross-tabulation ===")
        ct = pd.crosstab(df['ai_norm'], df['ver_norm'], margins=True)
        print(ct)
        
        disputed = df[df['ai_norm'] != df['ver_norm']]
        print(f"\nTotal Disputed: {len(disputed)}")
        
        false_rework = disputed[
            (disputed['ai_norm'] == 'rework') & (disputed['ver_norm'] == 'approved')
        ]
        false_approve = disputed[
            (disputed['ai_norm'] == 'approved') & (disputed['ver_norm'].isin(['rework']))
        ]
        print(f"False Rework (AI=Rework, Verifier=Approved): {len(false_rework)}")
        print(f"False Approve (AI=Approved, Verifier=Rework): {len(false_approve)}")
        
        # Parameter failure analysis
        param_cols = [c for c in df.columns if c and 'Met' in str(c)]
        print(f"\nParameter columns: {param_cols}")
        
        print("\n=== Parameter Failures in ALL Rework-by-AI Cases ===")
        ai_rework = df[df['ai_norm'] == 'rework']
        for p in param_cols:
            no_count = (ai_rework[p].astype(str).str.strip().str.lower() == 'no').sum()
            total = len(ai_rework)
            pct = no_count/total*100 if total > 0 else 0
            print(f"  {p}: {no_count}/{total} ({pct:.1f}%)")
        
        print("\n=== Parameter Failures in FALSE REWORK Cases ===")
        for p in param_cols:
            no_count = (false_rework[p].astype(str).str.strip().str.lower() == 'no').sum()
            total = len(false_rework)
            pct = no_count/total*100 if total > 0 else 0
            print(f"  {p}: {no_count}/{total} ({pct:.1f}%)")
        
        # Show consent reasons for false rework cases
        print("\n=== Consent Taken Reasons (False Rework) ===")
        consent_reason_col = [c for c in df.columns if c and 'Consent' in str(c) and 'Reason' in str(c)]
        if consent_reason_col:
            for i, (_, row) in enumerate(false_rework.iterrows()):
                fname = str(row.get('Filename', 'Unknown'))[:50]
                reason = str(row.get(consent_reason_col[0], ''))[:200]
                print(f"  Case {i+1} ({fname}): {reason}")
        
        # Show charges reasons for false rework cases
        print("\n=== Charges Explained Reasons (False Rework) ===")
        charges_reason_col = [c for c in df.columns if c and 'Charges' in str(c) and 'Reason' in str(c)]
        if charges_reason_col:
            for i, (_, row) in enumerate(false_rework.iterrows()):
                fname = str(row.get('Filename', 'Unknown'))[:50]
                reason = str(row.get(charges_reason_col[0], ''))[:200]
                print(f"  Case {i+1} ({fname}): {reason}")
        
        # Summary stats
        print("\n\n=== SUMMARY STATISTICS ===")
        total = len(df)
        agree = len(df[df['ai_norm'] == df['ver_norm']])
        print(f"Total Cases: {total}")
        print(f"Agreement: {agree} ({agree/total*100:.1f}%)")
        print(f"Disagreement: {total-agree} ({(total-agree)/total*100:.1f}%)")
        print(f"AI Approval Rate: {(df['ai_norm'] == 'approved').sum()}/{total} ({(df['ai_norm'] == 'approved').sum()/total*100:.1f}%)")
        print(f"Verifier Approval Rate: {(df['ver_norm'] == 'approved').sum()}/{total} ({(df['ver_norm'] == 'approved').sum()/total*100:.1f}%)")
