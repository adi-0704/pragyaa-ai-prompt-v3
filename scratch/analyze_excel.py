import pandas as pd
import json

# Read the Excel file - try different header rows since the overview screenshot shows merged cells
xlsx_path = 'AI vs Verifier audits DC.xlsx'

# First, read with openpyxl to understand structure better
import openpyxl
wb = openpyxl.load_workbook(xlsx_path, data_only=True)
print("Sheet names:", wb.sheetnames)

for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"\n=== Sheet: {sn} ({ws.max_row} rows x {ws.max_column} cols) ===")
    
    # Get all column headers
    headers = []
    for idx, cell in enumerate(ws[1]):
        v = cell.value if cell.value else f"Col_{idx+1}"
        headers.append(v)
    print("Headers:", headers)
    
    # Try to read as dataframe
    df = pd.read_excel(xlsx_path, sheet_name=sn)
    print(f"\nDataframe shape: {df.shape}")
    print(f"Columns: {list(df.columns)}")
    
    # Check if the headers match expected columns from openpyxl output
    if 'Call Status AI' in df.columns or any('Call Status' in str(c) for c in df.columns):
        ai_col = [c for c in df.columns if 'Call Status AI' in str(c)]
        ver_col = [c for c in df.columns if 'Call Status Verifier' in str(c) or 'Verifier' in str(c)]
        print(f"\nAI Status column: {ai_col}")
        print(f"Verifier Status column: {ver_col}")
    
    # Show unique values for status-like columns
    for c in df.columns:
        uniq = df[c].dropna().unique()
        if len(uniq) <= 10:
            print(f"\n  {c}: {list(uniq)}")

print("\n\n=== DEEP ANALYSIS ===")
# Re-read with proper column detection
for sn in wb.sheetnames:
    df = pd.read_excel(xlsx_path, sheet_name=sn)
    
    # Find AI and Verifier status columns
    ai_col = None
    ver_col = None
    for c in df.columns:
        cs = str(c).lower()
        if 'call status ai' in cs or 'status ai' in cs:
            ai_col = c
        if 'call status verifier' in cs or 'status verifier' in cs:
            ver_col = c
    
    if not ai_col or not ver_col:
        # Try by position based on what we saw from openpyxl
        print(f"Could not find AI/Verifier columns by name in sheet {sn}")
        print("Trying positional approach...")
        # Column K (11) = Call Status AI, Column L (12) = Call Status Verifier
        if len(df.columns) >= 12:
            ai_col = df.columns[10]  # 0-indexed
            ver_col = df.columns[11]
            print(f"Using positional: AI={ai_col}, Verifier={ver_col}")
    
    if ai_col and ver_col:
        print(f"\nAI Status: {ai_col}")
        print(f"Verifier Status: {ver_col}")
        
        # Normalize
        df['ai_norm'] = df[ai_col].astype(str).str.strip().str.lower()
        df['ver_norm'] = df[ver_col].astype(str).str.strip().str.lower()
        
        # Cross-tab
        print("\n=== Cross-tabulation ===")
        ct = pd.crosstab(df['ai_norm'], df['ver_norm'], margins=True)
        print(ct)
        
        # Disputed cases
        disputed = df[df['ai_norm'] != df['ver_norm']]
        print(f"\nTotal Disputed: {len(disputed)}")
        
        # False Reworks (AI=Rework, Verifier=Approved)
        false_rework = disputed[
            (disputed['ai_norm'] == 'rework') & (disputed['ver_norm'] == 'approved')
        ]
        print(f"False Rework (AI=Rework, Verifier=Approved): {len(false_rework)}")
        
        if len(false_rework) > 0:
            print("\n=== Parameter Failures in False Rework Cases ===")
            param_cols = [c for c in df.columns if 'Met' in str(c)]
            for p in param_cols:
                no_count = (false_rework[p].astype(str).str.strip().str.lower() == 'no').sum()
                total = len(false_rework)
                pct = no_count/total*100 if total > 0 else 0
                print(f"  {p}: {no_count}/{total} failures ({pct:.1f}%)")
            
            print("\n=== Detailed Reasons for False Rework Cases ===")
            reason_cols = [c for c in df.columns if 'Reason' in str(c)]
            for _, row in false_rework.iterrows():
                fname = str(row.get('Filename', 'Unknown'))[:60]
                print(f"\n--- Case: {fname} ---")
                print(f"  AI: {row[ai_col]} | Verifier: {row[ver_col]}")
                for rc in reason_cols:
                    val = str(row.get(rc, ''))
                    if val and val != 'nan':
                        print(f"  {rc}: {val[:120]}")
                # Show which params failed
                for p in param_cols:
                    if str(row.get(p, '')).strip().lower() == 'no':
                        reason_col = None
                        # Find corresponding reason column
                        for rc in reason_cols:
                            if p.replace(' Met', '') in rc:
                                reason_col = rc
                                break
                        reason = str(row.get(reason_col, ''))[:200] if reason_col else ''
                        print(f"  FAILED: {p} -> {reason}")
        
        # False Approve (AI=Approved, Verifier=Rework)  
        false_approve = disputed[
            (disputed['ai_norm'] == 'approved') & (disputed['ver_norm'] == 'rework')
        ]
        print(f"\nFalse Approve (AI=Approved, Verifier=Rework): {len(false_approve)}")
        
        if len(false_approve) > 0:
            print("\n=== False Approve Details ===")
            for _, row in false_approve.iterrows():
                fname = str(row.get('Filename', 'Unknown'))[:60]
                print(f"\n--- Case: {fname} ---")
                print(f"  AI: {row[ai_col]} | Verifier: {row[ver_col]}")
                reason_val = str(row.get('Reason for Rework', ''))
                if reason_val and reason_val != 'nan':
                    print(f"  Reason for Rework: {reason_val[:200]}")
                call_reason = str(row.get('Call Status Reasons', ''))
                if call_reason and call_reason != 'nan':
                    print(f"  Call Status Reasons: {call_reason[:200]}")
